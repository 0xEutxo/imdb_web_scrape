from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'top rated movies'
print(excel.sheetnames)
sheet.append(['Movie Rank','Title','Year Of Release', 'IMDB Rating'])
url = 'https://www.imdb.com/chart/top/?ref_=nv_mv_250'
source = requests.get(url)

soup = BeautifulSoup(source.text, 'html.parser')

movies = soup.find('tbody', class_='lister-list').find_all('tr')
for movie in movies:
    name = movie.find('td', class_='titleColumn').a.text
    rank = movie.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]
    year = movie.find('td', class_="titleColumn").span.text.strip('()')
    rating = movie.find('td', class_='ratingColumn imdbRating').strong.text
    print(rank, name, year, rating)
    sheet.append([rank, name, year, rating])

    excel.save('IMDB_Movie_Ratings.xlsx')
